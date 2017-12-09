# -*- coding: utf-8 -*-
"""
Created on Wed Oct 25 11:56:48 2017

@author: muhin
"""

import pymssql
import openpyxl
from openpyxl.comments import Comment

q = '"'
startDate = "2017-02-01"
endDate = "2017-03-01"
groupName = []
pID = []
intervalNo = []
tgID = []
tgSizeMPM = []
readWb = openpyxl.load_workbook("G:\\PycharmProjects\\InvitationCount\\readData.xlsx")
readSheet = readWb.get_sheet_by_name("groupName")
first_col = readSheet['A']

updateWb = openpyxl.load_workbook("G:\\PycharmProjects\\InvitationCount\\update.xlsx")

conn = pymssql.connect("localhost", "sa", "admin", "mds_results")
c1 = conn.cursor()

for x in range(len(first_col)):
    groupName.append(first_col[x].value)
    workSheet = updateWb.get_sheet_by_name(groupName[x])
    print("--- Group Name " + groupName[x] + " ---")
    # ******************** 1st Query ********************#
    c1.execute("""
               SELECT DISTINCT 
                P.[GroupName],
                R.[Project_ID],
                P.Name AS ProjectName,
                R.[Targetgroup_ID],
                R.[Interval_No],
                R.[DSCount],
                PTG.[InvitationLogicType],
                R.[TGCountAll],
                R.[TGCountUnique],
                R.[SampleCount],
                R.[SampleType],
                PTG.[Status]
                FROM RespondentLog R, Projects P, Projects_Targetgroups PTG
                WHERE len (R.ProcessingLog) > 0 AND R.[CreationDate] >= %s AND R.[CreationDate] < %s
                AND P.ID = R.Project_ID
                AND P.ID = PTG.Project_ID
                AND PTG.Status = 2
                AND P.GroupName = %s
                ORDER BY P.GroupName, R.Project_ID, R.Interval_No, R.Targetgroup_ID""",
               (startDate, endDate, groupName[x]))

    c1_list = c1.fetchall()
    length = len(c1_list)
    # print(length)
    r = 4
    c = 1  # c = 2

    for i in range(0, length):  # AND pt.Status = 2
        c1.execute("""
               SELECT Project_ID, ds.id DS_ID, T.id TG_ID,T.name TG_Name,pt.status as TG_Status, T.LastTotalSize TG_LastTotalSize
               FROM Targetgroups T, Projects_Targetgroups pt, DirectoryServers ds
               WHERE T.id= pt.Targetgroup_ID
               AND ds.id = T.DirServer_ID
               AND Project_ID = %d
               AND T.ID = %d
               AND T.DirServer_ID in (SELECT att.DirServer_ID FROM  DirServerAttributes att WHERE InvitationDistributor = 2)
               ORDER BY TG_Name,Project_ID""", (c1_list[i][1], c1_list[i][3]))
        c3_list = c1.fetchall()
        length3 = len(c3_list)
        # Friday Dec 8
        # print(length3)
        if length3 > 0:
            # print("This is distribution")
            c1.execute("""
                      SELECT R.Project_ID,R.Targetgroup_ID, R.Interval_No, SUM (R.[EmailToSend]) as 'ExpMailSent'
                      FROM RespondentLog R, Projects P
                      WHERE len (R.ProcessingLog) > 0
                      AND Project_ID = %d 
                      AND R.Interval_No = %d
                      AND P.ID = R.Project_ID
                      GROUP BY R.Interval_No, R.Targetgroup_ID, R.Project_ID
                      ORDER BY R.Interval_No""", (c1_list[i][1], c1_list[i][4]))
            c4_list = c1.fetchall()
            # print("ExpMailDist: " + str(c4_list[0][3]))

        # else:
        # print("This is not distribution")
        # print("i: " + str(i))
        tgID.append(c1_list[i][3])  # TG IDs
        tgIDasString = str(tgID[i])
        intervalNoZero = c1_list[i][4]
        if tgIDasString == "None" or intervalNoZero == 0:
            # print("This is NULL")
            pID.append(c1_list[i][1])
            intervalNo.append(c1_list[i][4])
            tgSizeMPM.append(c1_list[i][8])
            continue
        else:
            c1.execute("""
                      SELECT Project_ID, IntervalNo, IntervalName
                      FROM Intervals
                      where Project_ID = %d
                      AND IntervalNo = %d""", (c1_list[i][1], c1_list[i][4]))
            c5_list = c1.fetchall()
            workSheet.cell(row=r, column=c).value = c5_list[0][2]  # Interval Name
            # print("This is interval: " + c5_list[0][2])

            c = c + 1
            workSheet.cell(row=r, column=c).value = c1_list[i][1]  # project ID
            # print("Project id: " + str(c1_list[i][1])) # Project ID
            pID.append(c1_list[i][1])
            c = c + 1
            workSheet.cell(row=r, column=c).value = (str(c1_list[i][2]))  # project name
            c = c + 1
            workSheet.cell(row=r, column=c).value = c1_list[i][4]  # interval No.
            intervalNo.append(c1_list[i][4])
            # print(intervalNo[i])
            c = c + 2
            workSheet.cell(row=r, column=c).value = c1_list[i][5]  # DS Size
            c = c + 2

            invLogicType = c1_list[i][6]
            if invLogicType == 2:
                tgType = "Hard"
            else:
                tgType = "Soft"

            workSheet.cell(row=r, column=c).value = tgType  # TG Type, if 2 then Hard Soft otherwise
            c = c + 1
            workSheet.cell(row=r, column=c).value = c1_list[i][7]
            c = c + 1
            workSheet.cell(row=r, column=c).value = c1_list[i][8]  # TG size as in MPM
            tgSizeMPM.append(c1_list[i][8])
            c = c + 1
            if c1_list[i][10] == "%":
                sampleTypeP = str(c1_list[i][9])
                printSampleTypeP = sampleTypeP.replace('.0', '')
                # print(printSampleTypeP)
                workSheet.cell(row=r, column=c).value = printSampleTypeP + str(c1_list[i][10])  # Sample type %
            elif c1_list[i][10] == "All":
                workSheet.cell(row=r, column=c).value = str(c1_list[i][10])  # Sample type All
            elif c1_list[i][10] == "Target":
                workSheet.cell(row=r, column=c).value = str(c1_list[i][10])  # Sample type Target
            elif c1_list[i][10] == "Fixed":
                sampleTypeF = str(c1_list[i][9])
                printSampleTypeF = sampleTypeF.replace('.0', '')
                # print("Sample type:" + printSampleType)
                workSheet.cell(row=r, column=c).value = "Fixed(" + printSampleTypeF + ")"  # Sample type Fixed
            # Formulas
            c = c + 3
            workSheet.cell(row=r, column=c).value = "=M" + str(r) + "-L" + str(r)
            c = c + 1
            workSheet.cell(row=r, column=c).value = "=(N" + str(r) + "/L" + str(r) + ")"
            c = c + 1
            workSheet.cell(row=r, column=c).value = "=IF(O" + str(r) + "<>" + q + q + ",ABS(O" + str(
                r) + ")," + q + q + ")"

            c = 5
            # ******************** 2nd Query ********************#
            # print("pID: " + str(pID[i]))
            # print("interval: " + str(intervalNo[i]))
            # print("tgID: " + str(tgID[i]))
            c1.execute("""
                       SELECT D.Name as 'DS Name', T.Name as 'TG Name', R.Interval_No, T.ID, count (*) as 'Actual Mail Sent'
                       FROM RespondentLog R, Targetgroups T, DirectoryServers D
                       WHERE Project_ID = %d   
                       AND R.Interval_No = %d
                       AND T.id = R.Targetgroup_ID
                       AND T.id = %d
                       AND T.DirServer_ID = D.id
                       GROUP BY R.Project_ID, R.Interval_No, T.ID, T.Name, D.Name
                       ORDER BY R.Project_ID, R.Interval_No""", (pID[i], intervalNo[i], tgID[i]))
            c2_list = c1.fetchall()
            # len2 = len(c2_list)
            # print(len2)
            # for j in range(0,len2):
            print(c2_list[0][0])  # DS Name
            workSheet.cell(row=r, column=c).value = c2_list[0][0]  # DS name
            c = c + 2
            workSheet.cell(row=r, column=c).value = c2_list[0][1]  # TG name
            c = c + 5  # Expected (Calculated as per Inv. Logic)
            if c1_list[i][10] == "%" and length3 == 0:
                workSheet.cell(row=r, column=c).value = "=TRUNC(J" + str(r) + "*K" + str(
                    r) + ")"  # Formula =TRUNC(J4*K4)
            elif c1_list[i][10] == "All" and length3 == 0:
                workSheet.cell(row=r, column=c).value = tgSizeMPM[i]  # L = J

            elif c1_list[i][10] == "Target" and length3 == 0:
                workSheet.cell(row=r, column=c).value = tgSizeMPM[i] * 1.25  # L = K * 1.25
            elif c1_list[i][10] == "Fixed" and length3 == 0:
                workSheet.cell(row=r, column=c).value = tgSizeMPM[i]  # L = J
            elif c1_list[i][10] != "" and length3 > 0:
                workSheet.cell(row=r, column=c).value = c4_list[0][3]  # Distribution
                comment = Comment('Distribution enabled', 'MetatudeAsia')
                workSheet.cell(row=r, column=c).comment = comment
            c = c + 1
            workSheet.cell(row=r, column=c).value = c2_list[0][4]  # Actual Mail Sent
            c = 1
            r = r + 1
    pID.clear()
    intervalNo.clear()
    tgID.clear()
    c1_list.clear()
    c2_list.clear()
    c5_list.clear()
    tgSizeMPM.clear()
updateWb.save('update.xlsx')
print("All Done... :D")
conn.close()
