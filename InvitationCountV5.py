# -*- coding: utf-8 -*-
"""
Created on Wed Oct 25 11:56:48 2017

@author: muhin
"""

import sys
import pymssql
import openpyxl
from openpyxl.comments import Comment
from StringProcessing import processIntervalName, adhocIntervalName, fileReName

print('Enter Start & End date to start Invitation Count Monitoring for that month.\nUse this format only (yyyy-mm-dd)')
print('Start Date: ')
startDate = input()
print('End Date: ')
endDate = input()
print('Enter path of excel file to read customer names: ')
readData = input()
print('Enter path of excel file which is used as template: ')
updateData = input()
q = '"'
# startDate = "2017-02-01"  # yyyy-mm-dd
# endDate = "2017-03-01"
groupName = []
pID = []
intervalNo = []
tgID = []
tgSizeMPM = []
c1_list = []
c2_list = []
c5_list = []
readWb = openpyxl.load_workbook(readData)   # ('G:\\PycharmProjects\\InvitationCount\\readData.xlsx')
readSheet = readWb['groupName']
first_col = readSheet['A']

updateWb = openpyxl.load_workbook(updateData)  # ('G:\\PycharmProjects\\InvitationCount\\update.xlsx')

conn = pymssql.connect('localhost', 'sa', 'admin', 'mds_results')
c1 = conn.cursor()


for x in range(len(first_col)):
    groupName.append(first_col[x].value)
    workSheet = updateWb[groupName[x]]
    print("--- Group Name " + groupName[x] + " ---")
    # ******************** 1st Query ******************** #
    c1.execute("""
               SELECT DISTINCT 
                P.[GroupName],
                R.[Project_ID],
                P.Name AS ProjectName,
                R.[Targetgroup_ID],
                R.[Interval_No],
                R.[DSCount],
                R.[TGCountAll],
                R.[TGCountUnique],
                R.[SampleCount],
                R.[SampleType],
                P.[ProjectType]
                FROM RespondentLog R, Projects P
                WHERE len (R.ProcessingLog) > 0 AND R.[CreationDate] >= %s AND R.[CreationDate] < %s
                AND P.ID = R.Project_ID
                AND P.GroupName = %s
                 and R.[TGCountUnique] in (
                      select max ([TGCountUnique]) FROM RespondentLog R, Projects P
                      where len (R.ProcessingLog) > 0 AND R.[CreationDate] >= %s AND R.[CreationDate] < %s
                      AND P.ID = R.Project_ID
                      GROUP BY R.Targetgroup_ID, R.[Interval_No]
                      )
                ORDER BY P.GroupName, R.Project_ID, R.Interval_No, R.Targetgroup_ID""",
               (startDate, endDate, groupName[x], startDate, endDate))

    c1_list = c1.fetchall()
    length = len(c1_list)
    # print(length)
    r = 4
    c = 1  # c = 2

    for i in range(0, length):  # AND pt.Status = 2
        # print("Project Type: " + str(c1_list[i][10]))
        # Query for Distribution checking
        c1.execute("""
               SELECT Project_ID, ds.id DS_ID, T.id TG_ID, T.name TG_Name, pt.status TG_Status, T.LastTotalSize TG_LastTotalSize,
               att.Name att_name, Threshold
               FROM Targetgroups T, Projects_Targetgroups pt, DirectoryServers ds, DirServerAttributes att
               WHERE T.id= pt.Targetgroup_ID
               AND ds.id = T.DirServer_ID
               AND Project_ID = %d
               AND T.ID = %d
               AND T.DirServer_ID = att.DirServer_ID
               AND att.InvitationDistributor = 2
               ORDER BY Project_ID, TG_Name""", (c1_list[i][1], c1_list[i][3]))
        c3_list = c1.fetchall()
        # print(c3_list[0][6])
        # print(c3_list[0][7])
        length3 = len(c3_list)
        # Sunday Dec 10 ******************************************************************
        c1.execute("""
                SELECT [ID],[Project_ID],[Targetgroup_ID],[OptionType],[SampleSize],[Status],[InvitationLogicType] 
                FROM Projects_Targetgroups
                WHERE [Status] in(2,4)
                AND [Project_ID] = %d
                AND [Targetgroup_ID] = %d""", (c1_list[i][1], c1_list[i][3]))
        c6_list = c1.fetchall()
        # print(length3)
        if length3 > 0:
            # print("This is distribution")
            c1.execute("""
                      SELECT R.Project_ID, R.Targetgroup_ID, R.Interval_No, SUM(R.[EmailToSend]) as 'ExpMailSent'
                      FROM RespondentLog R, Projects P
                      WHERE len (R.ProcessingLog) > 0
                      AND Project_ID = %d 
                      AND R.Interval_No = %d
                      AND [Targetgroup_ID] = %d
                      AND P.ID = R.Project_ID
                      GROUP BY R.Interval_No, R.Targetgroup_ID, R.Project_ID
                      ORDER BY R.Interval_No""", (c1_list[i][1], c1_list[i][4], c1_list[i][3]))
            c4_list = c1.fetchall()
            # print("ExpMailDist: " + str(c4_list[0][3]))

        # else:
        # print("This is not distribution")
        # print("i: " + str(i))
        tgID.append(c1_list[i][3])  # TG IDs
        tgIDasString = str(tgID[i])
        intervalNoZero = c1_list[i][4]
        if tgIDasString == "None":  # or intervalNoZero == 0:
            # print("This is NULL")
            pID.append(c1_list[i][1])
            intervalNo.append(c1_list[i][4])
            tgSizeMPM.append(c1_list[i][7])  # c1_list[i][8])*********************************************************
            continue
        else:
            c1.execute("""
                      SELECT Project_ID, IntervalNo, IntervalName
                      FROM Intervals
                      where Project_ID = %d
                      AND IntervalNo = %d""", (c1_list[i][1], c1_list[i][4]))
            c5_list = c1.fetchall()
            # workSheet.cell(row=r, column=c).value = c5_list[0][2]  # Interval Name
            # Interval Name Enhancement *****************************************************************************
            if c1_list[i][10] == 1:
                # intervalName = "NA"
                intervalName = adhocIntervalName(startDate)
            else:
                intervalName = processIntervalName(c5_list[0][2])
            workSheet.cell(row=r, column=c).value = intervalName  # Interval Name
            # print("This is interval: " + c5_list[0][2])

            c = c + 1
            workSheet.cell(row=r, column=c).value = c1_list[i][1]  # project ID
            # print("Project id: " + str(c1_list[i][1])) # Project ID
            pID.append(c1_list[i][1])
            c = c + 1
            workSheet.cell(row=r, column=c).value = (str(c1_list[i][2]))  # project name
            c = c + 1
            if c1_list[i][4] == 0:
                intervalNumber = "NA"
            else:
                intervalNumber = c1_list[i][4]
            workSheet.cell(row=r, column=c).value = intervalNumber  # c1_list[i][4]  # interval No.
            intervalNo.append(c1_list[i][4])
            # print(intervalNo[i])
            c = c + 2
            workSheet.cell(row=r, column=c).value = c1_list[i][5]  # DS Size
            c = c + 2
            # print(i)
            invLogicType = c6_list[0][6]  # c1_list[i][6] InvitationLogicType *********************
            if invLogicType == 1:
                tgType = "Soft"
            elif invLogicType == 2:
                tgType = "Hard"
            else:
                tgType = "NA"
            workSheet.cell(row=r, column=c).value = tgType  # TG Type, if 1 = soft, 2 = hard
            c = c + 1
            workSheet.cell(row=r, column=c).value = c1_list[i][6]  # c1_list[i][7] ******************
            c = c + 1
            workSheet.cell(row=r, column=c).value = c1_list[i][7]  # TG size as in MPM
            tgSizeMPM.append(c1_list[i][7])
            c = c + 1
            if c1_list[i][9] == "%":
                sampleTypeP = str(c1_list[i][8])
                printSampleTypeP = sampleTypeP.replace('.0', '')
                # print(printSampleTypeP)
                workSheet.cell(row=r, column=c).value = printSampleTypeP + str(c1_list[i][9])  # Sample type %
            elif c1_list[i][9] == "All":
                workSheet.cell(row=r, column=c).value = str(c1_list[i][9])  # Sample type All
            elif c1_list[i][9] == "Target":
                workSheet.cell(row=r, column=c).value = str(c1_list[i][9])  # Sample type Target
            elif c1_list[i][9] == "Fixed":
                sampleTypeF = str(c1_list[i][8])
                printSampleTypeF = sampleTypeF.replace('.0', '')
                # print("Sample type:" + printSampleType)
                workSheet.cell(row=r, column=c).value = "Fixed(" + printSampleTypeF + ")"  # Sample type Fixed
            # Formulas
            c = c + 3
            workSheet.cell(row=r, column=c).value = "=M" + str(r) + "-L" + str(r)
            c = c + 1
            workSheet.cell(row=r, column=c).value = "=(N" + str(r) + "/L" + str(r) + ")"
            c = c + 1
            workSheet.cell(row=r, column=c).value = "=IF(O" + str(r) + "<>" + q + q + ",ABS(O" + str(
                r) + ")," + q + q + ")"

            c = 5
            # ******************** 2nd Query ********************#
            # print("pID: " + str(pID[i]))
            # print("interval: " + str(intervalNo[i]))
            # print("tgID: " + str(tgID[i]))
            c1.execute("""
                       SELECT D.Name as 'DS Name', T.Name as 'TG Name', R.Interval_No, T.ID, count (*) as 'Actual Mail Sent'
                       FROM RespondentLog R, Targetgroups T, DirectoryServers D
                       WHERE Project_ID = %d   
                       AND R.Interval_No = %d
                       AND T.id = R.Targetgroup_ID
                       AND T.id = %d
                       AND T.DirServer_ID = D.id
                       AND R.EmailCounter > -1
                       GROUP BY R.Project_ID, R.Interval_No, T.ID, T.Name, D.Name
                       ORDER BY R.Project_ID, R.Interval_No""", (pID[i], intervalNo[i], tgID[i]))
            c2_list = c1.fetchall()
            # len2 = len(c2_list)
            # print(len2)
            # for j in range(0,len2):
            print(c2_list[0][0])  # DS Name
            workSheet.cell(row=r, column=c).value = c2_list[0][0]  # DS name
            c = c + 2
            workSheet.cell(row=r, column=c).value = c2_list[0][1]  # TG name
            c = c + 5  # Expected (Calculated as per Inv. Logic)
            if c1_list[i][9] == "%" and length3 == 0:
                workSheet.cell(row=r, column=c).value = "=TRUNC(J" + str(r) + "*K" + str(
                    r) + ")"  # Formula =TRUNC(J4*K4)
            elif c1_list[i][9] == "All" and length3 == 0:
                workSheet.cell(row=r, column=c).value = tgSizeMPM[i]  # L = J

            elif c1_list[i][9] == "Target" and length3 == 0:
                workSheet.cell(row=r, column=c).value = tgSizeMPM[i] * 1.25  # L = K * 1.25
            elif c1_list[i][9] == "Fixed" and length3 == 0:
                workSheet.cell(row=r, column=c).value = c1_list[i][8]  # L = K
            elif c1_list[i][9] != "" and length3 > 0:
                workSheet.cell(row=r, column=c).value = c4_list[0][3]  # Distribution
                # print("Expected: " + str(c4_list[0][3]))
                attributeName = c3_list[0][6]
                threshold = str(c3_list[0][7])
                comment = Comment('Inv. Distributor- ' + attributeName + '\n' + 'Threshold- ' + threshold,
                                  'Metatude@sia')
                workSheet.cell(row=r, column=c).comment = comment
            c = c + 1
            workSheet.cell(row=r, column=c).value = c2_list[0][4]  # Actual Mail Sent
            c = 1
            r = r + 1
    pID.clear()
    intervalNo.clear()
    tgID.clear()
    c1_list.clear()
    c2_list.clear()
    c5_list.clear()
    tgSizeMPM.clear()
fileName = fileReName(startDate)
updateWb.save('Report on Invitation Count_' + fileName + '.xlsx')
# updateWb.save('update.xlsx')
print('All Done... :)')
print('Please find the report named: Report on Invitation Count_' + fileName + '.xlsx')
conn.close()
print('Press any key to Exit')
temp = input()
sys.exit(0)
